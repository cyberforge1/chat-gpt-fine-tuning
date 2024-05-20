import openpyxl
import requests
import os
import openai
from dotenv import load_dotenv

load_dotenv()

api_key = os.getenv("OPENAI_API_KEY")

def chat_with_gpt3(prompt):
    try:
        response = openai.ChatCompletion.create(
            model="gpt-3.5-turbo",  # Using the GPT-3.5 Turbo model
            messages=[
                {"role": "user", "content": prompt + " provide the type of responses to be returned from chatGPT for the question dataset"}
            ],
            max_tokens=150  # Adjust this value to control the response length
        )
        print(response['choices'][0]['message']['content'].strip())
        return response['choices'][0]['message']['content'].strip()
    except Exception as e:
        return str(e)


def process_excel_file(file_name):
    current_directory = os.path.dirname(os.path.abspath(__file__))
    file_path = os.path.join(current_directory, file_name)
    
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    
    # Add 'Completions' to the first row of the second column
    sheet.cell(row=1, column=2, value='completion')

    for idx, row in enumerate(sheet.iter_rows(min_row=2, min_col=1, max_col=1, values_only=True), start=2):
        prompt = row[0]
        completion = chat_with_gpt3(prompt)
        sheet.cell(row=idx, column=2, value=completion)


    workbook.save(file_path)


file_name = '../data/promptData.xlsx'


process_excel_file(file_name)



